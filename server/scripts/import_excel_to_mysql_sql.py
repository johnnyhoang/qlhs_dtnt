#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Read 2 Excel files and generate a MySQL import SQL file for the schema:
nguoi_dung, phan_quyen, hoc_sinh, bao_hiem, suat_an,
dot_thanh_toan, khoan_thanh_toan, dinh_muc_xe, don_gia_xe

Usage:
  python import_excel_to_mysql_sql.py \
    --excel1 /path/to/file1.xlsx \
    --excel2 /path/to/file2.xlsx \
    --out output_import.sql

Assumptions:
- Sheet names match table names (default). You can override in SHEET_MAP.
- Each sheet has a header row at row 1.
- For FK tables referencing hoc_sinh, the sheet should include 'ma_hoc_sinh' (or mapped equivalent).
"""

import argparse
import datetime as dt
import decimal
import uuid
from typing import Any, Dict, List, Tuple, Optional

from openpyxl import load_workbook


# -----------------------------
# CONFIG YOU MAY EDIT
# -----------------------------

# If your sheet names differ from table names, map here:
# key = table name, value = sheet name in Excel
SHEET_MAP: Dict[str, str] = {
    # "hoc_sinh": "DS_HOC_SINH",
    # "bao_hiem": "BAO_HIEM",
}

# Column mapping per table:
# key: table name
# value: mapping from excel header -> db column name
# If a header already equals db column name, you can omit it.
COLUMN_MAP: Dict[str, Dict[str, str]] = {
    "hoc_sinh": {
        "Mã HS": "ma_hoc_sinh",
        "Mã học sinh": "ma_hoc_sinh",
        "Họ tên": "ho_ten",
        "Lớp": "lop",
        "Ngày sinh": "ngay_sinh",
        "Giới tính": "gioi_tinh",
        "Trạng thái": "trang_thai",
        "CCCD": "cccd",
        "Mã MOET": "ma_moet",
    },
    "bao_hiem": {
        "Mã HS": "ma_hoc_sinh",  # used for lookup -> hoc_sinh_id (not inserted)
        "Số thẻ": "so_the",
        "Nơi ĐK": "noi_dang_ky",
        "Hạn sử dụng": "han_su_dung",
        "Địa chỉ": "dia_chi",
        "Thông tin sai": "thong_tin_sai",
        "Thông tin đúng": "thong_tin_dung",
        "Đã nộp ảnh": "da_nop_anh",
        "Lý do ảnh": "ly_do_anh",
        "Tra cứu VSSID": "tra_cuu_vssid",
        "Ngày đủ 5 năm": "ngay_du_5_nam",
        "Ghi chú": "ghi_chu",
        "Mã đối tượng": "ma_doi_tuong",
    },
    "suat_an": {
        "Mã HS": "ma_hoc_sinh",  # lookup only
        "Ngày": "ngay",
        "Loại suất ăn": "loai_suat_an",
        "Báo cắt": "bao_cat",
        "Ghi chú": "ghi_chu",
    },
    "dinh_muc_xe": {
        "Mã HS": "ma_hoc_sinh",  # lookup only
        "Số tài khoản": "so_tai_khoan",
        "Ngân hàng": "ngan_hang",
        "Địa chỉ xã mới": "dia_chi_xa_moi",
        "Xã hưởng trợ cấp": "xa_huong_tro_cap",
        "Khoảng cách": "khoang_cach",
        "Tên điểm dừng": "ten_diem_dung",
        "Phương tiện": "phuong_tien",
    },
    "dot_thanh_toan": {
        "Tháng": "thang",
        "Năm": "nam",
        "Ghi chú": "ghi_chu",
    },
    "khoan_thanh_toan": {
        "Mã HS": "ma_hoc_sinh",  # lookup only
        "Tháng": "thang",       # used to find dot_thanh_toan_id
        "Năm": "nam",           # used to find dot_thanh_toan_id
        "Xã": "xa",
        "Tiền ăn": "tien_an",
        "Tiền xe": "tien_xe",
        "Hỗ trợ khác": "ho_tro_khac",
        "Tổng tiền": "tong_tien",
        "Trạng thái": "trang_thai",
        "Ngày chi trả": "ngay_chi_tra",
        "Ghi chú": "ghi_chu",
    },
    "don_gia_xe": {
        "Vùng cũ": "vung_cu",
        "Vùng mới": "vung_moi",
        "Mức hỗ trợ": "muc_ho_tro",
        "Ghi chú": "ghi_chu",
    },
    "nguoi_dung": {
        "Email": "email",
        "Họ tên": "ho_ten",
        "Ảnh đại diện": "anh_dai_dien",
        "Vai trò": "vai_tro",
        "Kích hoạt": "kich_hoat",
    },
    "phan_quyen": {
        "Email": "email",      # lookup -> nguoi_dung_id
        "Mã module": "ma_module",
        "Có quyền xem": "co_quyen_xem",
        "Có quyền sửa": "co_quyen_sua",
    },
}

# Insert order to satisfy FKs
TABLE_ORDER: List[str] = [
    "nguoi_dung",
    "phan_quyen",
    "hoc_sinh",
    "bao_hiem",
    "dinh_muc_xe",
    "don_gia_xe",
    "dot_thanh_toan",
    "suat_an",
    "khoan_thanh_toan",
]

# Which tables should use upsert
UPSERT_KEYS: Dict[str, List[str]] = {
    "nguoi_dung": ["email"],
    "hoc_sinh": ["ma_hoc_sinh"],
    # bao_hiem: usually 1 row per hoc_sinh; if you have UNIQUE(hoc_sinh_id) then upsert will work.
    # If not unique in DB, you can switch to plain INSERT.
    "suat_an": ["hoc_sinh_id", "ngay", "loai_suat_an"],
    "khoan_thanh_toan": ["dot_thanh_toan_id", "hoc_sinh_id"],
    # phan_quyen: if you have unique(nguoi_dung_id, ma_module) in DB, add it here.
}

# -----------------------------
# Helpers
# -----------------------------

def normalize_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()

def to_bool(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "có", "co", "x"):
        return 1
    if s in ("0", "false", "no", "n", "không", "khong"):
        return 0
    return None

def to_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    # try parse common formats dd/mm/yyyy, yyyy-mm-dd
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None

def to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, (float,)):
        return int(round(v))
    if isinstance(v, decimal.Decimal):
        return int(v)
    s = str(v).strip().replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return None

def to_decimal(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    try:
        d = decimal.Decimal(str(v).replace(",", "").strip())
        return f"{d:.2f}"
    except Exception:
        return None

def sql_escape(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, (int, float, decimal.Decimal)):
        return str(value)
    s = str(value)
    s = s.replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"

def pick_sheet_name(table: str) -> str:
    return SHEET_MAP.get(table, table)

def read_sheet_rows(xlsx_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(h) for h in rows[0]]
    data: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        obj = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            obj[h] = r[idx] if idx < len(r) else None
        data.append(obj)
    return data

def map_columns(table: str, row: Dict[str, Any]) -> Dict[str, Any]:
    m = COLUMN_MAP.get(table, {})
    out: Dict[str, Any] = {}
    for k, v in row.items():
        db_col = m.get(k, k)  # if not mapped, assume already db col
        out[db_col] = v
    return out


# -----------------------------
# Main transformation logic
# -----------------------------

def build_sql(excel1: str, excel2: str, out_sql: str) -> None:
    # Load all tables data (from both excels)
    raw_by_table: Dict[str, List[Dict[str, Any]]] = {t: [] for t in TABLE_ORDER}

    for table in TABLE_ORDER:
        sheet = pick_sheet_name(table)
        raw_by_table[table].extend(read_sheet_rows(excel1, sheet))
        raw_by_table[table].extend(read_sheet_rows(excel2, sheet))

    # Map columns
    data_by_table: Dict[str, List[Dict[str, Any]]] = {}
    for table, rows in raw_by_table.items():
        mapped = [map_columns(table, r) for r in rows]
        data_by_table[table] = mapped

    # Build lookup: hoc_sinh.ma_hoc_sinh -> hoc_sinh.id
    hoc_sinh_rows = data_by_table.get("hoc_sinh", [])
    hs_id_by_code: Dict[str, str] = {}

    for r in hoc_sinh_rows:
        code = (r.get("ma_hoc_sinh") or "").strip()
        if not code:
            continue
        if "id" in r and r["id"]:
            hs_id_by_code[code] = str(r["id"]).strip()
        else:
            hs_id_by_code[code] = str(uuid.uuid4())

    # Build lookup: nguoi_dung.email -> id (we cannot know auto IDs; we will generate SQL that resolves with subquery)
    # We'll handle phan_quyen by using (SELECT id FROM nguoi_dung WHERE email=...)
    # Build dot_thanh_toan lookup: (thang,nam) -> resolved via INSERT + SELECT. We'll store as helper temp table.
    # Approach: create a TEMP TABLE import_dot_map to map (thang,nam) -> id after upsert.

    lines: List[str] = []
    lines.append("SET NAMES utf8mb4;")
    lines.append("SET FOREIGN_KEY_CHECKS=0;")
    lines.append("START TRANSACTION;")
    lines.append("")
    lines.append("-- Helper map for dot_thanh_toan (thang,nam) -> id")
    lines.append("DROP TEMPORARY TABLE IF EXISTS import_dot_map;")
    lines.append("""
CREATE TEMPORARY TABLE import_dot_map (
  thang INT NOT NULL,
  nam INT NOT NULL,
  dot_id INT NULL,
  PRIMARY KEY (thang, nam)
) ENGINE=Memory;
""".strip())
    lines.append("")

    def emit_insert(table: str, row: Dict[str, Any], upsert: bool) -> None:
        cols = list(row.keys())
        vals = [row[c] for c in cols]

        col_sql = ", ".join(f"`{c}`" for c in cols)
        val_sql = ", ".join(sql_escape(v) for v in vals)

        if not upsert or table not in UPSERT_KEYS:
            lines.append(f"INSERT INTO `{table}` ({col_sql}) VALUES ({val_sql});")
            return

        # ON DUPLICATE KEY UPDATE all non-key columns
        keys = set(UPSERT_KEYS[table])
        upd_cols = [c for c in cols if c not in keys]
        if not upd_cols:
            lines.append(f"INSERT INTO `{table}` ({col_sql}) VALUES ({val_sql}) ON DUPLICATE KEY UPDATE `{cols[0]}`=`{cols[0]}`;")
            return
        upd_sql = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in upd_cols)
        lines.append(f"INSERT INTO `{table}` ({col_sql}) VALUES ({val_sql}) ON DUPLICATE KEY UPDATE {upd_sql};")

    # 1) nguoi_dung
    for r in data_by_table.get("nguoi_dung", []):
        row = {}
        for c in ("email", "ho_ten", "anh_dai_dien", "vai_tro", "kich_hoat"):
            if c in r:
                row[c] = r[c]
        if "kich_hoat" in row:
            b = to_bool(row["kich_hoat"])
            row["kich_hoat"] = 1 if b is None else b
        if "vai_tro" in row and row["vai_tro"]:
            row["vai_tro"] = str(row["vai_tro"]).strip().upper()
        if "email" not in row or not row["email"]:
            continue
        emit_insert("nguoi_dung", row, upsert=True)

    lines.append("")

    # 2) phan_quyen (resolve nguoi_dung_id by subquery)
    for r in data_by_table.get("phan_quyen", []):
        email = (r.get("email") or "").strip()
        if not email or not r.get("ma_module"):
            continue
        co_xem = to_bool(r.get("co_quyen_xem"))
        co_sua = to_bool(r.get("co_quyen_sua"))
        co_xem = 1 if co_xem is None else co_xem
        co_sua = 0 if co_sua is None else co_sua

        # Use INSERT ... SELECT
        lines.append(
            "INSERT INTO `phan_quyen` (`nguoi_dung_id`,`ma_module`,`co_quyen_xem`,`co_quyen_sua`)\n"
            f"SELECT id, {sql_escape(r.get('ma_module'))}, {co_xem}, {co_sua} FROM `nguoi_dung` WHERE email={sql_escape(email)}\n"
            "ON DUPLICATE KEY UPDATE `co_quyen_xem`=VALUES(`co_quyen_xem`), `co_quyen_sua`=VALUES(`co_quyen_sua`);"
        )

    lines.append("")

    # 3) hoc_sinh
    for r in hoc_sinh_rows:
        code = (r.get("ma_hoc_sinh") or "").strip()
        if not code:
            continue
        row = {
            "id": hs_id_by_code[code],
            "ma_hoc_sinh": code,
            "ho_ten": r.get("ho_ten"),
            "lop": r.get("lop"),
            "ma_moet": r.get("ma_moet"),
            "cccd": r.get("cccd"),
            "gioi_tinh": (str(r.get("gioi_tinh")).strip().upper() if r.get("gioi_tinh") else None),
            "trang_thai": (str(r.get("trang_thai")).strip().upper() if r.get("trang_thai") else None),
        }
        if r.get("ngay_sinh") is not None:
            row["ngay_sinh"] = to_date(r.get("ngay_sinh"))
        # defaults if missing
        if not row.get("gioi_tinh"):
            row["gioi_tinh"] = "KHAC"
        if not row.get("trang_thai"):
            row["trang_thai"] = "DANG_HOC"
        emit_insert("hoc_sinh", row, upsert=True)

    lines.append("")

    # Helper for FK resolve from ma_hoc_sinh
    def hs_id(code: Any) -> Optional[str]:
        c = (code or "").strip()
        return hs_id_by_code.get(c)

    # 4) bao_hiem
    for r in data_by_table.get("bao_hiem", []):
        hid = hs_id(r.get("ma_hoc_sinh") or r.get("ma_hoc_sinh".upper()))
        if not hid:
            continue
        row = {
            "hoc_sinh_id": hid,
            "ma_doi_tuong": r.get("ma_doi_tuong"),
            "so_the": r.get("so_the"),
            "noi_dang_ky": r.get("noi_dang_ky"),
            "han_su_dung": to_date(r.get("han_su_dung")),
            "dia_chi": r.get("dia_chi"),
            "thong_tin_sai": r.get("thong_tin_sai"),
            "thong_tin_dung": r.get("thong_tin_dung"),
            "da_nop_anh": (to_bool(r.get("da_nop_anh")) or 0),
            "ly_do_anh": r.get("ly_do_anh"),
            "tra_cuu_vssid": r.get("tra_cuu_vssid"),
            "ngay_du_5_nam": to_date(r.get("ngay_du_5_nam")),
            "ghi_chu": r.get("ghi_chu"),
        }
        emit_insert("bao_hiem", row, upsert=False)  # đổi True nếu DB có UNIQUE(hoc_sinh_id)

    lines.append("")

    # 5) dinh_muc_xe
    for r in data_by_table.get("dinh_muc_xe", []):
        hid = hs_id(r.get("ma_hoc_sinh"))
        if not hid:
            continue
        row = {
            "hoc_sinh_id": hid,
            "so_tai_khoan": r.get("so_tai_khoan"),
            "ngan_hang": r.get("ngan_hang"),
            "dia_chi_xa_moi": r.get("dia_chi_xa_moi"),
            "xa_huong_tro_cap": r.get("xa_huong_tro_cap"),
            "khoang_cach": to_decimal(r.get("khoang_cach")),
            "ten_diem_dung": r.get("ten_diem_dung"),
            "phuong_tien": r.get("phuong_tien"),
        }
        emit_insert("dinh_muc_xe", row, upsert=False)  # đổi True nếu DB có UNIQUE(hoc_sinh_id)

    lines.append("")

    # 6) don_gia_xe
    for r in data_by_table.get("don_gia_xe", []):
        if not r.get("vung_cu") and not r.get("vung_moi"):
            continue
        row = {
            "vung_cu": r.get("vung_cu"),
            "vung_moi": r.get("vung_moi"),
            "muc_ho_tro": to_int(r.get("muc_ho_tro")) or 0,
            "ghi_chu": r.get("ghi_chu"),
        }
        emit_insert("don_gia_xe", row, upsert=False)

    lines.append("")

    # 7) dot_thanh_toan + fill import_dot_map
    for r in data_by_table.get("dot_thanh_toan", []):
        thang = to_int(r.get("thang"))
        nam = to_int(r.get("nam"))
        if not thang or not nam:
            continue
        ghi_chu = r.get("ghi_chu")
        # Upsert dot_thanh_toan by (thang,nam) if you have UNIQUE(thang,nam). If not, it will insert duplicates.
        lines.append(
            "INSERT INTO `dot_thanh_toan` (`thang`,`nam`,`ghi_chu`)\n"
            f"VALUES ({thang},{nam},{sql_escape(ghi_chu)})\n"
            "ON DUPLICATE KEY UPDATE `ghi_chu`=VALUES(`ghi_chu`);"
        )
        lines.append(
            "REPLACE INTO import_dot_map (thang, nam, dot_id)\n"
            f"SELECT {thang}, {nam}, id FROM `dot_thanh_toan` WHERE thang={thang} AND nam={nam} ORDER BY id DESC LIMIT 1;"
        )

    lines.append("")

    # 8) suat_an
    for r in data_by_table.get("suat_an", []):
        hid = hs_id(r.get("ma_hoc_sinh"))
        if not hid:
            continue
        ngay = to_date(r.get("ngay"))
        loai = (str(r.get("loai_suat_an")).strip().upper() if r.get("loai_suat_an") else None)
        if not ngay or not loai:
            continue
        row = {
            "hoc_sinh_id": hid,
            "ngay": ngay,
            "loai_suat_an": loai,
            "bao_cat": (to_bool(r.get("bao_cat")) or 0),
            "ghi_chu": r.get("ghi_chu"),
        }
        emit_insert("suat_an", row, upsert=True)

    lines.append("")

    # 9) khoan_thanh_toan (resolve dot_thanh_toan_id by thang/nam from import_dot_map)
    for r in data_by_table.get("khoan_thanh_toan", []):
        hid = hs_id(r.get("ma_hoc_sinh"))
        thang = to_int(r.get("thang"))
        nam = to_int(r.get("nam"))
        if not hid or not thang or not nam:
            continue

        tien_an = to_int(r.get("tien_an")) or 0
        tien_xe = to_int(r.get("tien_xe")) or 0
        ho_tro_khac = to_int(r.get("ho_tro_khac")) or 0
        tong_tien = to_int(r.get("tong_tien"))
        if tong_tien is None:
            tong_tien = tien_an + tien_xe + ho_tro_khac

        trang_thai = (str(r.get("trang_thai")).strip().upper() if r.get("trang_thai") else "CHO_XU_LY")
        ngay_chi_tra = to_date(r.get("ngay_chi_tra"))

        # Insert via SELECT (dot_id from import_dot_map)
        lines.append(
            "INSERT INTO `khoan_thanh_toan`\n"
            "(`dot_thanh_toan_id`,`hoc_sinh_id`,`xa`,`tien_an`,`tien_xe`,`ho_tro_khac`,`tong_tien`,`trang_thai`,`ngay_chi_tra`,`ghi_chu`)\n"
            "SELECT m.dot_id, "
            f"{sql_escape(hid)}, "
            f"{sql_escape(r.get('xa'))}, "
            f"{tien_an}, {tien_xe}, {ho_tro_khac}, {tong_tien}, {sql_escape(trang_thai)}, {sql_escape(ngay_chi_tra)}, {sql_escape(r.get('ghi_chu'))}\n"
            f"FROM import_dot_map m WHERE m.thang={thang} AND m.nam={nam}\n"
            "ON DUPLICATE KEY UPDATE "
            "`xa`=VALUES(`xa`),"
            "`tien_an`=VALUES(`tien_an`),"
            "`tien_xe`=VALUES(`tien_xe`),"
            "`ho_tro_khac`=VALUES(`ho_tro_khac`),"
            "`tong_tien`=VALUES(`tong_tien`),"
            "`trang_thai`=VALUES(`trang_thai`),"
            "`ngay_chi_tra`=VALUES(`ngay_chi_tra`),"
            "`ghi_chu`=VALUES(`ghi_chu`);"
        )

    lines.append("")
    lines.append("COMMIT;")
    lines.append("SET FOREIGN_KEY_CHECKS=1;")
    lines.append("")

    with open(out_sql, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"✅ Wrote SQL import file: {out_sql}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel1", required=True, help="Path to excel file 1 (.xlsx)")
    ap.add_argument("--excel2", required=True, help="Path to excel file 2 (.xlsx)")
    ap.add_argument("--out", required=True, help="Output .sql path")
    args = ap.parse_args()

    build_sql(args.excel1, args.excel2, args.out)


if __name__ == "__main__":
    main()
